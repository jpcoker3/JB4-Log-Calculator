import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import math
import openpyxl.utils.dataframe as dataframe
import openpyxl 
from openpyxl.chart import ScatterChart,Reference, Series
from openpyxl.chart.trendline import Trendline

whole_list = []

def handle_graphs(calculations_file):
    
    global whole_list
    
    map0 = {"map" : "0",
            "rpm": [],
            "boost": [],
            "gear": [],
            "throttle": []}
    map1 = {"map" : "1",
            "rpm": [],
            "boost": [],
            "gear": [],
            "throttle": []}
    map2 = {"map" : "2", 
            "rpm": [],
            "boost": [],
            "gear": [],
            "throttle": []}
    map3 = {"map" : "3",
            "rpm": [],
            "boost": [],
            "gear": [],
            "throttle": []}
    map4 = {"map" : "4",
            "rpm": [],
            "boost": [],
            "gear": [],
            "throttle": []}
    map5 = {"map" : "5",
            "rpm": [],
            "boost": [],
            "gear": [],
            "throttle": []}
    map6 = {"map" : "6",
            "rpm": [],
            "boost": [],
            "gear": [],
            "throttle": []}
    map7 = {"map" : "7",
            "rpm": [],
            "boost": [],
            "gear": [],
            "throttle": []}
    map8 = {"map" : "8",
            "rpm": [],
            "boost": [],
            "gear": [],
            "throttle": []}
    list_of_dicts = [map0, map1, map2, map3, map4, map5, map6, map7, map8]
    
    
    for item in whole_list: #item is each file in Logs folder
        for i in range(item["map"].size): #goes throught each column in map
            #print(item["map"][i])
            pass
        
        #match the map of each file to the correct case. adds the rpm, boost, and gear to the correct map dictionary
        match item["map"][1]:
            case 0:
                for i in range(item["map"].size):
                    map0["rpm"].append(item["rpm"][i])
                    map0["boost"].append(item["Manifold Boost"][i])
                    map0["gear"].append(item["gear"][i])
                    map0["throttle"].append(item["throttle"][i])
            case 1:
                for i in range(item["map"].size):
                    map1["rpm"].append(item["rpm"][i])
                    map1["boost"].append(item["Manifold Boost"][i])
                    map1["gear"].append(item["gear"][i])
                    map1["throttle"].append(item["throttle"][i])
            case 2:
                for i in range(item["map"].size):
                    map2["rpm"].append(item["rpm"][i])
                    map2["boost"].append(item["Manifold Boost"][i])
                    map2["gear"].append(item["gear"][i])
                    map2["throttle"].append(item["throttle"][i])
            case 3:
                for i in range(item["map"].size):
                    map3["rpm"].append(item["rpm"][i])
                    map3["boost"].append(item["Manifold Boost"][i])
                    map3["gear"].append(item["gear"][i])
                    map3["throttle"].append(item["throttle"][i])
            case 4:
                for i in range(item["map"].size):
                    map4["rpm"].append(item["rpm"][i])
                    map4["boost"].append(item["Manifold Boost"][i])
                    map4["gear"].append(item["gear"][i])
                    map4["throttle"].append(item["throttle"][i])
            case 5:
                for i in range(item["map"].size):
                    map5["rpm"].append(item["rpm"][i])
                    map5["boost"].append(item["Manifold Boost"][i])
                    map5["gear"].append(item["gear"][i])
                    map5["throttle"].append(item["throttle"][i])
            case 6:
                for i in range(item["map"].size):
                    map6["rpm"].append(item["rpm"][i])
                    map6["boost"].append(item["Manifold Boost"][i])
                    map6["gear"].append(item["gear"][i])
                    map6["throttle"].append(item["throttle"][i])
            case 7:
                for i in range(item["map"].size):
                    map7["rpm"].append(item["rpm"][i])
                    map7["boost"].append(item["Manifold Boost"][i])
                    map7["gear"].append(item["gear"][i])
                    map7["throttle"].append(item["throttle"][i])
            case 8:
                for i in range(item["map"].size):
                    map8["rpm"].append(item["rpm"][i])
                    map8["boost"].append(item["Manifold Boost"][i])
                    map8["gear"].append(item["gear"][i])
                    map8["throttle"].append(item["throttle"][i])
    
    
    
    # we create the sheet here so that the graphs are created on a new sheet
    #if created in the loop, there are overwrite issues and this is just better. 
    wb = openpyxl.load_workbook(calculations_file)
    
    # Check if the "Graphs" sheet already exists and remove it
    if "Graphs" in wb.sheetnames:
        wb.remove(wb["Graphs"])
    
    # Create a new sheet named "Graphs"
    sheet = wb.create_sheet("Graphs")
    
    counter = 0 #this is used to determine the graph position
    #create the graphs
    for map in list_of_dicts:
        
        if map["rpm"] == []: #if the map is empty
            pass
        else: #if the map is not empty
            counter +=1 #increment the counter only if creating new graph
            create_graphs(map, calculations_file, sheet, wb, counter)
    return




def create_graphs(map_dictionary, calculations_file,sheet, wb, counter):
    End_of_last_row = 1

   
     
    rpms = []
    boosts = []
    gears = []
    throttle = []
    # Write the data from the map_dictionary to the "Graphs" sheet
    for i in range(len(map_dictionary["rpm"])):
        rpms.append(map_dictionary["rpm"][i])
        boosts.append(map_dictionary["boost"][i])
        gears.append(map_dictionary["gear"][i])
        throttle.append(map_dictionary["throttle"][i])
        
        #sheet.append([map_dictionary["rpm"][i], map_dictionary["boost"][i], map_dictionary["gear"][i]])
        
    
    # Create an object of ScatterChart class
    chart = ScatterChart()
    
    gear_dict = {
        1: {
                "rpm":[],
                "boost":[],
                "throttle": []
        },
        2: {    
                "rpm":[],
                "boost":[],
                "throttle": []
        },
        3: {
                "rpm":[],
                "boost":[],
                "throttle": []
        },
        4: {
                "rpm":[],
                "boost":[],
                "throttle": []
        },
        5: {
                "rpm":[],
                "boost":[],
                "throttle": []
        },
        6: {
                "rpm":[],
                "boost":[],
                "throttle": []
        },
        7: {
                "rpm":[],
                "boost":[],
                "throttle": []      
        },
        8: {
                "rpm":[],
                "boost":[],
                "throttle": []
        }
    }

    # Iterate through the unique gear values and create series for each gear
    unique_gears = set(map_dictionary["gear"])
    
    for gear in unique_gears:
        for i in range(len(gears)):
            if gears[i] == gear:
                gear_dict[gear]["rpm"].append(rpms[i])
                gear_dict[gear]["boost"].append(boosts[i])
                gear_dict[gear]["throttle"].append(throttle[i])
                
                

   
    for gear in unique_gears:
        new_max = 0
        to_add = []
        for i in range(len(gear_dict[gear]["rpm"])):
            if gear_dict[gear]["throttle"][i] > 50: #only add points where throttle is greater than 50%
                to_add.append([gear_dict[gear]["rpm"][i], gear_dict[gear]["boost"][i], gear])
        
        for row in to_add:
            sheet.append(row)
            new_max += 1
        
        # Update the references for each gear
        xvalues = Reference(sheet, min_col=1, min_row=End_of_last_row + 1, max_row=End_of_last_row + new_max, max_col=1)
        yvalues = Reference(sheet, min_col=2, min_row=End_of_last_row + 1, max_row=End_of_last_row + new_max, max_col=2)
        
        series = Series(yvalues, xvalues, title="Gear " + str(gear))

        # Append the series to the chart
        chart.series.append(series)

        End_of_last_row += new_max
        

    # Set the title of the chart
    chart.title = "Boost Per Gear (Map " + map_dictionary["map"] + ")"

    # Set the title of the x-axis
    chart.x_axis.title = "RPM"

    # Set the title of the y-axis
    chart.y_axis.title = "Manifold Boost"

    # Add the chart to the "Graphs" sheet
    sheet.add_chart(chart, "A" + str((counter * 16) - 15))


    # Save the file
    wb.save(calculations_file)
    

    return



def round_rpm(rpm):
    if math.isnan(rpm):
        return rpm
    return int(rpm // 500) * 500

def average_duplicate_rows(df, unique_cols):
    # Group the dataframe by unique columns and calculate the mean for other columns
    grouped = df.groupby(unique_cols).mean().round(2) #rounds to 2 decimal places
    return grouped

def create_averages(file_path, sheet_name, file_format):
    # Read the file based on the format (CSV or XLSX)
    if file_format == 'csv':
        df = pd.read_csv(file_path, header=4)
    elif file_format == 'xlsx':
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=4)
    else:
        raise ValueError("Invalid file format. Only 'csv' and 'xlsx' are supported." +
                         "\n Received: {}".format(file_format))

    # Convert the 'boost' and 'boost2' columns to numeric, skipping non-numeric values
    df['Chargepipe Boost'] = pd.to_numeric(df['boost'], errors='coerce')
    df['Manifold Boost'] = pd.to_numeric(df['boost2'], errors='coerce')

    # Round RPM values down to the nearest 500
    df['rounded_rpm'] = df['rpm'].apply(round_rpm)

    # Calculate the average values per rounded RPM per each gear
    averages_whole = df
    averages = df.groupby(['map', 'gear', 'rounded_rpm']).mean().round(2)

    # Select the desired columns
    columns = ['ecu_psi', 'Chargepipe Boost', 'Manifold Boost', 'throttle', 'afr', 'mph']
    averages = averages[columns]

    return [averages, averages_whole]

def create_calcs(folder_path, calculation_file):
    # Create an empty dictionary to store the files and sheet names
    files_and_sheet_names = {}

    # Iterate over the files in the folder
    for filename in os.listdir(folder_path):
        if filename.endswith((".xlsx", ".csv")):
            file_path = os.path.join(folder_path, filename)
            if filename.endswith(".xlsx"):
                file_format = "xlsx"
                wb = load_workbook(file_path, read_only=True)
                sheet_names = wb.sheetnames
            else:
                file_format = "csv"
                sheet_names = [None]  # For CSV files, use a placeholder sheet name
            files_and_sheet_names[file_path] = sheet_names

    # Create an empty list to store all the averages
    all_averages = []

    # Iterate over the dictionary items
    for excel_file, sheet_names in files_and_sheet_names.items():
        for sheet_name in sheet_names:
            # Get the file format from the file extension
            file_format = excel_file.split('.')[-1]

            # Get the averages for the current file and sheet
            return_ = create_averages(excel_file, sheet_name, file_format)
            global whole_list
            whole_list.append(return_[1]) #used for graphs. 
            # Append the averages to the overall list
            all_averages.append(return_[0])

    # Concatenate all the averages into a single DataFrame
    all_averages_df = pd.concat(all_averages)
    all_averages_df = average_duplicate_rows(all_averages_df, ['map', 'gear', 'rounded_rpm'])

    # Write the DataFrame to the calculation file
    with pd.ExcelWriter(calculation_file, mode='a', engine='openpyxl') as writer:
        # Check if the 'Averages' sheet already exists in the Excel file
        if 'Averages' in writer.book.sheetnames:
            # Delete the existing 'Averages' sheet
            writer.book.remove(writer.book['Averages'])
            writer.book.save(calculation_file)

        # Write the DataFrame to a new 'Averages' sheet
        all_averages_df.to_excel(writer, sheet_name='Averages')
        writer.book._sheets = [writer.book['Averages']] + [sheet for sheet in writer.book._sheets if sheet.title != 'Averages']

        # Formatting
        # Set the column widths and alignment
        # Color the afr column based on values
        worksheet = writer.sheets['Averages']
        for column in worksheet.columns:
            column_width = 15
            worksheet.column_dimensions[column[0].column_letter].width = column_width

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        writer.book.save(calculation_file)  # Save the workbook

    # Apply conditional formatting to the 'afr' column after saving the workbook
    book = load_workbook(calculation_file)
    sheet = book['Averages']

    column_location = 8  # 8 to account for the first 3 columns
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=column_location, max_col=column_location):
        for cell in row:
            
            #fill yellow by default
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            #fill green if between 12.5 and 13
            if cell.value is not None and 12.5 <= cell.value <= 13:
                fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                
            #fill red if less than 9.0 or greater than 15  
            elif cell.value is not None and (cell.value < 9.0 or cell.value > 15):
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.fill = fill

    # Save the workbook again to apply the conditional formatting
    book.save(calculation_file)

def main():
    #ex: calculation_file = r"C:\Path\To\Save\File.xlsx" 
    # #where you want to save the calculations. sheet must exist and have at least one sheet (default)
    calculation_file = r"C:\Users\jpcok\Documents\CarStuff\Tiguan\JB4\Calculations.xlsx"
    
    # Load the log path
    #ex: log_path = r"C:\Path\To\Log\Folder" #contains all the log files within (.csv or .xlsx)
    log_path = r"C:\Users\jpcok\Documents\CarStuff\Tiguan\JB4\Logs"

    create_calcs(log_path, calculation_file)
    handle_graphs(calculation_file)

if __name__ == "__main__":
    main()
